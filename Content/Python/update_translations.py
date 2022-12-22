# Copyright (C) 2022 Dakarai Simmons - All Rights Reserved

import os
import io
import re
import sys
import shutil

# The directory of the file (to search for other files)
# And a map of languages to codes (for .po files)
home_dir = os.path.dirname(__file__)
language_codes = {
    "English": "en",
    "Spanish (Latin America)": "es",
    "Spanish (Spain)": "es-ES",
    "Chinese": "zh-Hans",
    "Japanese": "ja-JP",
    "French": "fr",
    "Russian": "ru",
    "Polish": "pl",
    "German": "de",
    "Korean": "ko"
}

excel_path = ""
case_sensitive = False
force_refresh = False
pages = []
languages = []
translations = {}

# Open the file with settings and update the variables
with open(os.path.join(home_dir, "Temp", "update.txt"), "r") as params:
    excel_path = params.readline().strip(' \n')
    case_sensitive = (params.readline().strip(' \n') == "True")
    force_refresh = (params.readline().strip(' \n') == "True")
    sizes = re.findall("\d+", params.readline().strip(' \n'))
    page_num = int(sizes[0])
    lang_num = int(sizes[1])
    for i in range(page_num):
        pages.append(params.readline().strip(' \n'))
    for j in range(lang_num):
        languages.append(params.readline().strip(' \n'))

# Import the .egg libraries for excel reading
for path in os.listdir(home_dir):
	p = os.path.join(home_dir, path)
	ext = os.path.splitext(path)[1]
	if os.path.isfile(p) and ext == ".egg":
		sys.path.append(p) 

from openpyxl import load_workbook

# Escape double quotes so they don't mess with po file parsing
def escape_quotes(string):
    return string.replace("\"", "\\\"")

def escape_newlines(string):
    if(string.find("\r\n") != -1):
        return string.strip().replace("\r\n", "\\r\\n")
    else:
        return string.strip().replace("\n", "\\n")

def get_pairs(string, char):
    indexes = []
    i = 0
    count = 0
    pair = [0, 0]

    for s in string:
        if s == char:
            pair[count % 2] = i
            if (count % 2) == 1:
                indexes.append((pair[0], pair[1]))
            count += 1
        i += 1

    if (string.count(char) % 2) == 1:
        indexes.append((pair[0], 0))

    return indexes

# So the way this works is:
# It reads all the requested pages and saves the strings
# to a dictionary of english strings mapped to another dictionary of chosen 
# languages (Spanish, Korean, etc.) to the string value from that cell.
#{
# "Phrase 1": {"Spanish": "Spanish phrase", "Korean": "Korean phrase", ...},
# ...
# }
# For cells with full sentences (or multiple sentences), the spreadsheet has special 
# invisible characters (u\2060) sandwitching each phrase as they're spit up in game.
# The script will take the cell and separate them if the character is found.
# Once the dictionary is built, it searches through the exported .po files
# and maps a translation if found.
def update():
    wb = load_workbook(excel_path)
    for page in pages:
        worksheet = wb[page]
        lang_index = {}
        first_row = worksheet[1]
        for cell in first_row:
            if cell.value != "Keys" and cell.value != "English" and cell.value != None:
                lang_index[cell.column_letter] = cell.value
        for row in worksheet.iter_rows(min_row=2, min_col=2):
            english_key = []
            for cell in row:
                if cell.value != None and cell.column_letter == "B":
                    if cell.value.count(u'\u2060') > 0:
                        pairs = get_pairs(cell.value, u'\u2060')
                        for pair in pairs:
                            s = escape_quotes(cell.value[pair[0] + 1:pair[1]])
                            s = s if case_sensitive else s.lower()
                            english_key.append(s.strip())
                            translations[s.strip()] = {}
                    else:
                        s = cell.value if case_sensitive else cell.value.lower()
                        s = escape_quotes(s).strip()
                        english_key.append(s)
                        translations[s] = {}
                elif cell.value != None and lang_index.get(cell.column_letter) != None and languages.count(lang_index[cell.column_letter]) > 0:
                    if cell.value.count(u'\u2060') > 0:
                        pairs = get_pairs(escape_quotes(cell.value), u'\u2060')
                        len = english_key.__len__()
                        for i in range(len):
                            s = escape_quotes(cell.value)
                            translations[english_key[i]][lang_index[cell.column_letter]] = s[pairs[i][0] + 1:pairs[i][1]]
                    else:
                        translations[english_key[0]][lang_index[cell.column_letter]] = escape_quotes(cell.value)
    
    for lang in languages:
        # Note for this method: This is in Python 2 (because of what this version of Unreal ships with)
        # and to save Unicode text, you need to use io with the 'utf8' encoding, and have each
        # string literal be a unicode literal.
        with io.open(os.path.join(home_dir, "Temp", language_codes[lang] + ".po"), "w", encoding="utf8") as new_file:
            file_path = os.path.join(home_dir, "../../../../Content/Localization/Game/{}/Game.po".format(language_codes[lang]))
            with io.open(file_path, encoding="utf8") as old_file:
                msgid = ""
                for line in old_file:
                    if(line.find("msgid ") != -1 and line[7:-2] != ""):
                        msgid = line[7:-2].rstrip()
                        new_file.write(line)
                    elif(line.find("msgstr ") != -1 and msgid != ""):
                        if force_refresh or (line.strip() == "msgstr \"\""):
                            s = msgid if case_sensitive else msgid.lower()
                            dic = translations.get(s)
                            if dic == None:
                                new_file.write(u"msgstr \"\"\n")
                            else:
                                new_file.write(u"msgstr \"{}\"\n".format(escape_newlines(dic.get(lang, ""))))
                        else:
                            new_file.write(line)
                        msgid = ""
                    else:
                        new_file.write(line)
                
        shutil.move(os.path.join(home_dir, "Temp", language_codes[lang] + ".po"), os.path.join(home_dir, "../../../../Content/Localization/Game/{}/Game.po".format(language_codes[lang])))

                            

# Leaving this here to test the output by itself
#if __name__ == "__main__":
#    update()