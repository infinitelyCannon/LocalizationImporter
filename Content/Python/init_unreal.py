# Copyright (C) 2022 Dakarai Simmons - All Rights Reserved

import os
import sys
import unreal

home_dir = os.path.dirname(__file__)
chosen_path = ""

# The translation tool depends on this external library
# for reading excel sheets, so this adds every library (.egg)
# file to the system path for importing
for path in os.listdir(home_dir):
	p = os.path.join(home_dir, path)
	ext = os.path.splitext(path)[1]
	if os.path.isfile(p) and ext == ".egg":
		sys.path.append(p)
		print("Added Module: " + p)

from openpyxl import load_workbook

# A child of the Python Bridge class to have one globally accessible reference in C++
@unreal.uclass()
class PythonBridgeImpl(unreal.PythonBridge):	
	# Reads through every page and add its info to
	# the list if it has translations
	# in order to differenciate between two types (Pages & Languages)
	# all pages start out False (unchecked) and Languages are True
	@unreal.ufunction(override=True)
	def import_spreadsheet(self, path):
		chosen_path = path
		language_page = ""
		wb = load_workbook(chosen_path)
		pages = wb.sheetnames
		result = []
		for page in pages:
			sheet = wb[page]
			if sheet['A1'].value == "Keys" and sheet['B1'].value == "English":
				struct = unreal.UpdateTranslationsSettings()
				struct.title = page
				struct.checked = False
				result.append(struct)
				language_page = page
		sheet = wb[language_page]
		row = sheet[1]
		for cell in row:
			if cell.value != None and cell.value != "Keys" and cell.value != "English":
				struct = unreal.UpdateTranslationsSettings()
				struct.title = cell.value
				struct.checked = True
				result.append(struct)
		return result

	@unreal.ufunction(override=True)
	def update_selection(self, path, pages, languages, refresh, case_sensitive):
		# Writes all the chosen settings to a file for a standalone script to read:
		# path to the spreadsheet
		# is it case-sensitive
		# should force refresh
		# list of used pages
		# list of used languages
		try:
			os.mkdir(os.path.join(home_dir, "Temp"))
		except:
			pass

		with open(os.path.join(home_dir, "Temp", "update.txt"), "w") as update_file:
			update_file.write(path + "\n")
			update_file.write(str(case_sensitive) + "\n")
			update_file.write(str(refresh) + "\n")
			update_file.write(pages)
			update_file.write(languages)
